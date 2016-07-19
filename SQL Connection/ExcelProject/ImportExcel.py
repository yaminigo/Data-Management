'''
Created on Feb 4, 2016

@author: yaminigoyal
'''
import openpyxl
from openpyxl.cell import get_column_letter
wb = openpyxl.load_workbook('purchases.xlsx')
fd = open("insertFile.sql","w");
for sheetname in wb.get_sheet_names():
    sheet = wb.get_sheet_by_name(sheetname)
    maxcell = get_column_letter(sheet.max_column) + `sheet.max_row`
    maxvalue = sheet.max_column
    if (sheetname== "Product"):
        maxcell ='D'+ `sheet.max_row`
        maxvalue = 4
    for rowOfCellObjects in sheet['A2':maxcell]:
        str1 = "insert into " + sheetname + " values("
        i =1;
        for cellObj in rowOfCellObjects:
            if((sheetname == "Company" or sheetname == "Product") and i== 2):
                str1 += str(cellObj.value) +  ","
            elif (i == maxvalue):
                str1 += "\"" + str(cellObj.value) + "\")"
            else:
                str1 += "\"" + str(cellObj.value) + "\","
            i = i+1
        str1 = str1+";"+"\n"
        fd.write(str1);
            
            